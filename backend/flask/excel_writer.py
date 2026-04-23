"""Write the graded results back to an Excel workbook the user can download."""

from __future__ import annotations

import io
from typing import Dict, List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _header_style(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="174C8F")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _grade_fill(grade: str) -> str:
    g = grade.upper()
    if g.startswith("A"):
        return "C6EFCE"
    if g.startswith("B"):
        return "BDD7EE"
    if g.startswith("C"):
        return "FFE699"
    if g.startswith("D"):
        return "F8CBAD"
    return "F4CCCC"


def build_results_workbook(
    student_grades: Dict[str, Dict[str, object]],
    student_scores: Dict[str, Dict[str, float]],
    weights: Dict[str, float],
    chart_data: List[Dict[str, object]],
    stats: Dict[str, float],
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grades"

    division_names = list(weights.keys())
    headers = ["Reg No", "Weighted Total %"] + division_names + ["Grade"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        _header_style(cell)

    row_num = 2
    for reg in sorted(student_grades.keys()):
        entry = student_grades[reg]
        ws.cell(row=row_num, column=1, value=reg)
        ws.cell(row=row_num, column=2, value=entry.get("total_percent"))
        for j, div in enumerate(division_names, start=3):
            ws.cell(row=row_num, column=j, value=student_scores.get(reg, {}).get(div))
        grade = str(entry.get("grade", ""))
        gcell = ws.cell(row=row_num, column=len(headers), value=grade)
        gcell.fill = PatternFill("solid", fgColor=_grade_fill(grade))
        gcell.font = Font(bold=True)
        row_num += 1

    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18

    # Stats sheet
    s = wb.create_sheet("Summary")
    s.append(["Metric", "Value"])
    for k, v in stats.items():
        s.append([k, v])
    s.append([])
    s.append(["Grade", "Count"])
    for row in chart_data:
        s.append([row.get("grade"), row.get("count")])
    for c in range(1, 3):
        s.column_dimensions[get_column_letter(c)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
