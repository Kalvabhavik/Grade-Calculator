"""Read an uploaded Excel workbook and extract per-student division scores."""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

import openpyxl

from matching import (
    _normalize,
    detect_header_row,
    detect_max_marks_row,
    detect_weight_row,
    match_columns,
)

REG_NO_ALIASES = {"regno", "registrationnumber", "regnumber", "rollno",
                   "rollnumber", "registernumber", "studentid", "id"}
NAME_ALIASES = {"name", "studentname", "fullname"}
TEAMID_ALIASES = {"teamid", "team", "groupid", "group"}
MEMBERS_ALIASES = {"membersregnos", "members", "teammembers", "memberregnos", "studentregnos"}


@dataclass
class SheetData:
    sheet_name: str
    is_team_sheet: bool
    headers: List[str]
    header_idx: int
    max_marks: Dict[str, float] = field(default_factory=dict)
    weights: Dict[str, float] = field(default_factory=dict)
    rows: List[Dict[str, object]] = field(default_factory=list)
    reg_no_col: Optional[str] = None
    name_col: Optional[str] = None
    team_id_col: Optional[str] = None
    members_col: Optional[str] = None
    division_cols: List[str] = field(default_factory=list)


def _load_workbook(path: str):
    return openpyxl.load_workbook(path, data_only=True)


def _rows_of(ws) -> List[List[object]]:
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _classify_header(header: str) -> Optional[str]:
    n = _normalize(header)
    if n in REG_NO_ALIASES:
        return "reg_no"
    if n in NAME_ALIASES:
        return "name"
    if n in TEAMID_ALIASES:
        return "team_id"
    if n in MEMBERS_ALIASES:
        return "members"
    return None


def parse_sheet(ws, division_names: List[str]) -> SheetData:
    rows = _rows_of(ws)
    if not rows:
        return SheetData(sheet_name=ws.title, is_team_sheet=False, headers=[], header_idx=0)

    header_idx = detect_header_row(rows)
    headers_raw = rows[header_idx]
    headers = [str(h).strip() if h is not None else "" for h in headers_raw]

    # classify the meta columns
    reg_no_col = name_col = team_id_col = members_col = None
    for h in headers:
        kind = _classify_header(h)
        if kind == "reg_no" and reg_no_col is None:
            reg_no_col = h
        elif kind == "name" and name_col is None:
            name_col = h
        elif kind == "team_id" and team_id_col is None:
            team_id_col = h
        elif kind == "members" and members_col is None:
            members_col = h

    meta_cols = {c for c in [reg_no_col, name_col, team_id_col, members_col] if c}
    division_headers = [h for h in headers if h and h not in meta_cols]

    is_team_sheet = team_id_col is not None or members_col is not None

    # optional rows
    max_row_idx = detect_max_marks_row(rows, header_idx)
    weight_row_idx = detect_weight_row(rows, header_idx)

    max_marks: Dict[str, float] = {}
    weights: Dict[str, float] = {}
    skip_rows = {header_idx}
    if max_row_idx is not None:
        skip_rows.add(max_row_idx)
        mrow = rows[max_row_idx]
        for i, h in enumerate(headers):
            if h in division_headers and i < len(mrow) and isinstance(mrow[i], (int, float)):
                max_marks[h] = float(mrow[i])
    if weight_row_idx is not None:
        skip_rows.add(weight_row_idx)
        wrow = rows[weight_row_idx]
        for i, h in enumerate(headers):
            if h in division_headers and i < len(wrow) and isinstance(wrow[i], (int, float)):
                weights[h] = float(wrow[i])

    # map divisions → actual headers in the sheet (fuzzy)
    mapping, _ = match_columns(division_headers, division_names)

    student_rows: List[Dict[str, object]] = []
    header_by_idx = {i: h for i, h in enumerate(headers)}
    for ri, row in enumerate(rows):
        if ri <= header_idx or ri in skip_rows:
            continue
        # empty row?
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue
        d: Dict[str, object] = {}
        for i, c in enumerate(row):
            h = header_by_idx.get(i)
            if not h:
                continue
            d[h] = c
        student_rows.append(d)

    return SheetData(
        sheet_name=ws.title,
        is_team_sheet=is_team_sheet,
        headers=headers,
        header_idx=header_idx,
        max_marks=max_marks,
        weights=weights,
        rows=student_rows,
        reg_no_col=reg_no_col,
        name_col=name_col,
        team_id_col=team_id_col,
        members_col=members_col,
        division_cols=division_headers,
    )


def parse_workbooks(paths: List[str], division_names: List[str]) -> List[SheetData]:
    out: List[SheetData] = []
    for p in paths:
        wb = _load_workbook(p)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row <= 1 and ws.max_column <= 1:
                continue
            out.append(parse_sheet(ws, division_names))
    return out


def _split_members(text: str) -> List[str]:
    if not text:
        return []
    parts = [p.strip() for p in str(text).replace(";", ",").replace("/", ",").split(",")]
    return [p for p in parts if p]


def build_student_scores(
    sheets: List[SheetData],
    division_names: List[str],
    mapping_override: Optional[Dict[str, Dict[str, str]]] = None,
) -> Tuple[Dict[str, Dict[str, float]], Dict[str, float], Dict[str, Dict[str, str]]]:
    """Merge per-sheet rows into per-student percentage scores per division.

    Returns
    -------
    student_scores : {reg_no: {division: percent_0_to_100}}
    max_marks      : {division: max_marks}
    mapping_used   : {sheet_name: {division: header}}
    """
    student_scores: Dict[str, Dict[str, float]] = {}
    max_marks_all: Dict[str, float] = {}
    mapping_used: Dict[str, Dict[str, str]] = {}

    # Pre-compute individual-sheet mappings (for team member fanout)
    individual_regs: Dict[str, str] = {}  # normalized reg → canonical reg
    for s in sheets:
        if s.is_team_sheet or not s.reg_no_col:
            continue
        for row in s.rows:
            reg = str(row.get(s.reg_no_col, "")).strip()
            if not reg:
                continue
            individual_regs[_normalize(reg)] = reg

    for s in sheets:
        # Mapping: override > fuzzy match
        if mapping_override and s.sheet_name in mapping_override:
            local_map = {d: h for d, h in mapping_override[s.sheet_name].items() if h}
        else:
            local_map, _ = match_columns(s.division_cols, division_names)
            local_map = {d: h for d, h in local_map.items() if h}
        mapping_used[s.sheet_name] = local_map

        # Accumulate per-division max marks (prefer highest if multiple sheets declare)
        for div, header in local_map.items():
            mx = s.max_marks.get(header)
            if mx is not None and mx > max_marks_all.get(div, 0):
                max_marks_all[div] = mx

        if s.is_team_sheet:
            for row in s.rows:
                members = _split_members(str(row.get(s.members_col, "") or ""))
                if not members:
                    continue
                resolved_members: List[str] = []
                for m in members:
                    canon = individual_regs.get(_normalize(m), m)
                    resolved_members.append(canon)
                for div, header in local_map.items():
                    raw = row.get(header)
                    if not isinstance(raw, (int, float)):
                        continue
                    mx = s.max_marks.get(header)
                    percent = (float(raw) / mx * 100.0) if mx and mx > 0 else float(raw)
                    for reg in resolved_members:
                        student_scores.setdefault(reg, {})[div] = percent
        else:
            if not s.reg_no_col:
                continue
            for row in s.rows:
                reg = str(row.get(s.reg_no_col, "")).strip()
                if not reg:
                    continue
                for div, header in local_map.items():
                    raw = row.get(header)
                    if not isinstance(raw, (int, float)):
                        continue
                    mx = s.max_marks.get(header)
                    percent = (float(raw) / mx * 100.0) if mx and mx > 0 else float(raw)
                    student_scores.setdefault(reg, {})[div] = percent

    return student_scores, max_marks_all, mapping_used


def preview_workbook(paths: List[str], division_names: List[str]) -> Dict[str, object]:
    """Return a JSON-friendly summary for the AI-preview screen."""
    sheets = parse_workbooks(paths, division_names)
    preview: List[Dict[str, object]] = []
    for s in sheets:
        mapping, scores = match_columns(s.division_cols, division_names)
        preview.append({
            "sheet_name": s.sheet_name,
            "is_team_sheet": s.is_team_sheet,
            "reg_no_column": s.reg_no_col,
            "name_column": s.name_col,
            "team_id_column": s.team_id_col,
            "members_column": s.members_col,
            "division_columns": s.division_cols,
            "row_count": len(s.rows),
            "max_marks": s.max_marks,
            "weights": s.weights,
            "mapping": mapping,
            "confidence": scores,
        })
    return {"sheets": preview}
